import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import json

def exportar_json_dashboard():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    
    archivo = filedialog.askopenfilename(
        title="Selecciona el archivo Excel de Portadas",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    root.destroy()

    if not archivo:
        return

    try:
        # 1. Usar openpyxl con data_only=True para leer VALORES de fórmulas
        # Nota: Si el excel no se ha guardado con los valores calculados,
        # esto leerá vacío. Asegúrate de abrir y guardar el Excel antes.
        import openpyxl
        wb = openpyxl.load_workbook(archivo, data_only=True)
        
        # Intentar tomar 'Hoja1', si no, la activa
        if 'Hoja1' in wb.sheetnames:
            ws = wb['Hoja1']
        else:
            ws = wb.active
            
        # Leer datos en una lista de diccionarios
        data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(row)]
            else:
                if not any(row): continue # Saltar filas totalmente vacías
                row_dict = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                data.append(row_dict)
        
        df = pd.DataFrame(data)

        # 2. Limpieza
        df = df.dropna(subset=['Titular'])
        
        campos_criticos = ['Grupo', 'Dependencia', 'Medio', 'Tema', 'Estatus', 'Recurso editorial']
        for col in campos_criticos:
            if col in df.columns:
                # Convertir a string para evitar errores de tipo
                df[col] = df[col].astype(str).fillna('Sin especificar').replace('None', 'Sin especificar')
            else:
                df[col] = 'Sin especificar'

        # 3. Formatear Fecha
        if 'Fecha' in df.columns:
            df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
            # Crear columna auxiliar para rango de fechas si es necesario, 
            # pero para JSON exportamos string ISO
            df['Fecha'] = df['Fecha'].dt.strftime('%Y-%m-%d')
            df['Fecha'] = df['Fecha'].fillna('Sin fecha')

        # 4. Exportar
        output_path = os.path.join(os.path.dirname(archivo), "data_portadas.json")
        data_dict = df.to_dict(orient='records')
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data_dict, f, ensure_ascii=False, indent=4)

        print(f"Éxito: Exportados {len(df)} registros.")
        messagebox.showinfo("Éxito", f"Archivo generado con {len(df)} registros.")

    except Exception as e:
        print(f"Error crítico: {e}")
        messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    exportar_json_dashboard()