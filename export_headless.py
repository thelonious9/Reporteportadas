import pandas as pd
import os
import json
import openpyxl

def run_headless_export():
    archivo = r'c:\Users\vivie\Reporteportadas\MuestraPortadas.xlsx'
    if not os.path.exists(archivo):
        print(f"Error: No se encontró {archivo}")
        return

    try:
        print(f"Leyendo {archivo}...")
        wb = openpyxl.load_workbook(archivo, data_only=True)
        if 'Hoja1' in wb.sheetnames:
            ws = wb['Hoja1']
        else:
            ws = wb.active
            
        data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(row)]
            else:
                if not any(row): continue
                row_dict = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                data.append(row_dict)
        
        df = pd.DataFrame(data)
        df = df.dropna(subset=['Titular'])
        
        campos_criticos = ['Grupo', 'Dependencia', 'Medio', 'Tema', 'Estatus', 'Recurso editorial', 'Nivel']
        for col in campos_criticos:
            if col in df.columns:
                df[col] = df[col].astype(str).fillna('Sin especificar').replace('None', 'Sin especificar')
            else:
                df[col] = 'Sin especificar'

        if 'Fecha' in df.columns:
            df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
            df['Fecha'] = df['Fecha'].dt.strftime('%Y-%m-%d')
            df['Fecha'] = df['Fecha'].fillna('Sin fecha')

        output_path = r'c:\Users\vivie\Reporteportadas\data_portadas.json'
        data_dict = df.to_dict(orient='records')
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data_dict, f, ensure_ascii=False, indent=4)

        print(f"Éxito: Exportados {len(df)} registros a {output_path}")

    except Exception as e:
        print(f"Error crítico: {e}")

if __name__ == "__main__":
    run_headless_export()
